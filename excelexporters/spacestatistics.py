import base64
import uuid
import os
from openpyxl.chart import (
    BarChart,
    Reference,
)
from openpyxl.styles import PatternFill, Border, Side, Alignment, Font
from openpyxl.drawing.image import Image
from openpyxl import Workbook
from openpyxl.chart.label import DataLabelList


####################################################################################################################
# PROCEDURES
# Step 1: Validate the report data
# Step 2: Generate excel file
# Step 3: Encode the excel file bytes to Base64
####################################################################################################################


def export(report,
           name,
           reporting_start_datetime_local,
           reporting_end_datetime_local,
           period_type):
    ####################################################################################################################
    # Step 1: Validate the report data
    ####################################################################################################################
    if report is None:
        return None
    print(report)

    ####################################################################################################################
    # Step 2: Generate excel file from the report data
    ####################################################################################################################
    filename = generate_excel(report,
                              name,
                              reporting_start_datetime_local,
                              reporting_end_datetime_local,
                              period_type)
    ####################################################################################################################
    # Step 3: Encode the excel file to Base64
    ####################################################################################################################
    try:
        with open(filename, 'rb') as binary_file:
            binary_file_data = binary_file.read()
    except IOError as ex:
        pass

    # Base64 encode the bytes
    base64_encoded_data = base64.b64encode(binary_file_data)
    # get the Base64 encoded data using human-readable characters.
    base64_message = base64_encoded_data.decode('utf-8')
    # delete the file from server
    try:
        os.remove(filename)
    except NotImplementedError as ex:
        pass
    return base64_message


def generate_excel(report,
                   name,
                   reporting_start_datetime_local,
                   reporting_end_datetime_local,
                   period_type):
    wb = Workbook()
    ws = wb.active

    # Row height
    ws.row_dimensions[1].height = 121

    for i in range(2, 37 + 1):
        ws.row_dimensions[i].height = 30

    for i in range(38, 90 + 1):
        ws.row_dimensions[i].height = 30

    # Col width
    ws.column_dimensions['A'].width = 1.5
    ws.column_dimensions['B'].width = 20.0

    for i in range(ord('C'), ord('I')):
        ws.column_dimensions[chr(i)].width = 15.0

    # Font
    name_font = Font(name='Constantia', size=15, bold=True)
    title_font = Font(name='宋体', size=15, bold=True)
    # data_font = Font(name='Franklin Gothic Book', size=11)

    table_fill = PatternFill(fill_type='solid', fgColor='1F497D')
    f_border = Border(left=Side(border_style='medium', color='00000000'),
                      right=Side(border_style='medium', color='00000000'),
                      bottom=Side(border_style='medium', color='00000000'),
                      top=Side(border_style='medium', color='00000000')
                      )
    b_border = Border(
        bottom=Side(border_style='medium', color='00000000'),
    )

    b_c_alignment = Alignment(vertical='bottom',
                              horizontal='center',
                              text_rotation=0,
                              wrap_text=False,
                              shrink_to_fit=False,
                              indent=0)
    c_c_alignment = Alignment(vertical='center',
                              horizontal='center',
                              text_rotation=0,
                              wrap_text=False,
                              shrink_to_fit=False,
                              indent=0)
    b_r_alignment = Alignment(vertical='bottom',
                              horizontal='right',
                              text_rotation=0,
                              wrap_text=False,
                              shrink_to_fit=False,
                              indent=0)
    # c_r_alignment = Alignment(vertical='bottom',
    #                           horizontal='center',
    #                           text_rotation=0,
    #                           wrap_text=False,
    #                           shrink_to_fit=False,
    #                           indent=0)

    # Img
    img = Image("excelexporters/myems.png")
    # img = Image("myems.png")
    ws.add_image(img, 'B1')

    # Title
    ws['B3'].font = name_font
    ws['B3'].alignment = b_r_alignment
    ws['B3'] = 'Name:'
    ws['C3'].border = b_border
    ws['C3'].alignment = b_c_alignment
    ws['C3'].font = name_font
    ws['C3'] = name

    ws['D3'].font = name_font
    ws['D3'].alignment = b_r_alignment
    ws['D3'] = 'Period:'
    ws['E3'].border = b_border
    ws['E3'].alignment = b_c_alignment
    ws['E3'].font = name_font
    ws['E3'] = period_type

    ws['F3'].font = name_font
    ws['F3'].alignment = b_r_alignment
    ws['F3'] = 'Date:'
    ws['G3'].border = b_border
    ws['G3'].alignment = b_c_alignment
    ws['G3'].font = name_font
    ws['G3'] = reporting_start_datetime_local + "__" + reporting_end_datetime_local
    ws.merge_cells("G3:H3")

    if "reporting_period" not in report.keys() or \
            "names" not in report['reporting_period'].keys() or len(report['reporting_period']['names']) == 0:
        filename = str(uuid.uuid4()) + '.xlsx'
        wb.save(filename)

        return filename
    #################################################
    # First: 统计分析
    # 6: title
    # 7: table title
    # 8~11 table_data
    # Total: 6 rows
    # if has not energy data: set low height for rows
    #################################################
    reporting_period_data = report['reporting_period']

    has_energy_data_flag = True

    if "names" not in reporting_period_data.keys() or \
            reporting_period_data['names'] is None or \
            len(reporting_period_data['names']) == 0:
        has_energy_data_flag = False

        filename = str(uuid.uuid4()) + '.xlsx'
        wb.save(filename)

        return filename

    if has_energy_data_flag:
        ws['B6'].font = title_font
        ws['B6'] = name + ' 统计分析'
        # ws['D6'].font = title_font
        # ws['D6'] = '面积' +report['space']['area']

        category = reporting_period_data['names']

        # table_title
        ws['B7'].fill = table_fill
        ws['B7'].font = title_font
        ws['B7'].alignment = c_c_alignment
        ws['B7'] = '报告期'
        ws['B7'].border = f_border

        ws['C7'].font = title_font
        ws['C7'].alignment = c_c_alignment
        ws['C7'] = '算术平均数'
        ws['C7'].border = f_border

        ws['D7'].font = title_font
        ws['D7'].alignment = c_c_alignment
        ws['D7'] = '中位数'
        ws['D7'].border = f_border

        ws['E7'].font = title_font
        ws['E7'].alignment = c_c_alignment
        ws['E7'] = '最小值'
        ws['E7'].border = f_border

        ws['F7'].font = title_font
        ws['F7'].alignment = c_c_alignment
        ws['F7'] = '最大值'
        ws['F7'].border = f_border

        ws['G7'].font = title_font
        ws['G7'].alignment = c_c_alignment
        ws['G7'] = '样本标准差'
        ws['G7'].border = f_border

        ws['H7'].font = title_font
        ws['H7'].alignment = c_c_alignment
        ws['H7'] = '样本方差'
        ws['H7'].border = f_border

        # table_data

        for i, value in enumerate(category):
            row = i*2 + 8
            ws['B' + str(row)].font = name_font
            ws['B' + str(row)].alignment = c_c_alignment
            ws['B' + str(row)] = reporting_period_data['names'][i] + " (" + reporting_period_data['units'][i] + " )"
            ws['B' + str(row)].border = f_border

            ws['B' + str(row + 1)].font = name_font
            ws['B' + str(row + 1)].alignment = c_c_alignment
            ws['B' + str(row + 1)] = "环比"
            ws['B' + str(row + 1)].border = f_border

            ws['C' + str(row)].font = name_font
            ws['C' + str(row)].alignment = c_c_alignment
            ws['C' + str(row)] = round(reporting_period_data['means'][i], 2) \
                if reporting_period_data['means'][i] is not None else ''
            ws['C' + str(row)].border = f_border

            ws['C' + str(row + 1)].font = name_font
            ws['C' + str(row + 1)].alignment = c_c_alignment
            ws['C' + str(row + 1)] = str(round(reporting_period_data['means_increment_rate'][i] * 100, 2)) + "%" \
                if reporting_period_data['means_increment_rate'][i] is not None else '0.00%'
            ws['C' + str(row + 1)].border = f_border

            ws['D' + str(row)].font = name_font
            ws['D' + str(row)].alignment = c_c_alignment
            ws['D' + str(row)] = round(reporting_period_data['medians'][i], 2) \
                if reporting_period_data['medians'][i] is not None else ''
            ws['D' + str(row)].border = f_border

            ws['D' + str(row + 1)].font = name_font
            ws['D' + str(row + 1)].alignment = c_c_alignment
            ws['D' + str(row + 1)] = str(round(reporting_period_data['medians_increment_rate'][i] * 100, 2)) + "%" \
                if reporting_period_data['medians_increment_rate'][i] is not None else '0.00%'
            ws['D' + str(row + 1)].border = f_border

            ws['E' + str(row)].font = name_font
            ws['E' + str(row)].alignment = c_c_alignment
            ws['E' + str(row)] = round(reporting_period_data['minimums'][i], 2) \
                if reporting_period_data['minimums'][i] is not None else ''
            ws['E' + str(row)].border = f_border

            ws['E' + str(row + 1)].font = name_font
            ws['E' + str(row + 1)].alignment = c_c_alignment
            ws['E' + str(row + 1)] = str(round(reporting_period_data['minimums_increment_rate'][i] * 100, 2)) + "%" \
                if reporting_period_data['minimums_increment_rate'][i] is not None else '0.00%'
            ws['E' + str(row + 1)].border = f_border

            ws['F' + str(row)].font = name_font
            ws['F' + str(row)].alignment = c_c_alignment
            ws['F' + str(row)] = round(reporting_period_data['maximums'][i], 2) \
                if reporting_period_data['maximums'][i] is not None else ''
            ws['F' + str(row)].border = f_border

            ws['F' + str(row + 1)].font = name_font
            ws['F' + str(row + 1)].alignment = c_c_alignment
            ws['F' + str(row + 1)] = str(round(reporting_period_data['maximums_increment_rate'][i] * 100, 2)) + "%" \
                if reporting_period_data['maximums_increment_rate'][i] is not None else '0.00%'
            ws['F' + str(row + 1)].border = f_border

            ws['G' + str(row)].font = name_font
            ws['G' + str(row)].alignment = c_c_alignment
            ws['G' + str(row)] = round(reporting_period_data['stdevs'][i], 2) \
                if reporting_period_data['stdevs'][i] is not None else ''
            ws['G' + str(row)].border = f_border

            ws['G' + str(row + 1)].font = name_font
            ws['G' + str(row + 1)].alignment = c_c_alignment
            ws['G' + str(row + 1)] = str(round(reporting_period_data['stdevs_increment_rate'][i] * 100, 2)) + "%" \
                if reporting_period_data['stdevs_increment_rate'][i] is not None else '0.00%'
            ws['G' + str(row + 1)].border = f_border

            ws['H' + str(row)].font = name_font
            ws['H' + str(row)].alignment = c_c_alignment
            ws['H' + str(row)] = round(reporting_period_data['variances'][i], 2) \
                if reporting_period_data['variances'][i] is not None else ''
            ws['H' + str(row)].border = f_border

            ws['H' + str(row + 1)].font = name_font
            ws['H' + str(row + 1)].alignment = c_c_alignment
            ws['H' + str(row + 1)] = str(round(reporting_period_data['variances_increment_rate'][i] * 100, 2)) + "%" \
                if reporting_period_data['variances_increment_rate'][i] is not None else '0.00%'
            ws['H' + str(row + 1)].border = f_border
    #################################################
    # First: 统计分析
    # 13: title
    # 14: table title
    # 15~18 table_data
    # Total: 6 rows
    # if has not energy data: set low height for rows
    #################################################

    if has_energy_data_flag:
        ws['B13'].font = title_font
        ws['B13'] = name + ' 单位面积值'
        ws['D13'].font = title_font
        ws['D13'] = str(report['space']['area']) + 'M²'

        category = reporting_period_data['names']

        # table_title
        ws['B14'].fill = table_fill
        ws['B14'].font = title_font
        ws['B14'].alignment = c_c_alignment
        ws['B14'] = '报告期'
        ws['B14'].border = f_border

        ws['C14'].font = title_font
        ws['C14'].alignment = c_c_alignment
        ws['C14'] = '算术平均数'
        ws['C14'].border = f_border

        ws['D14'].font = title_font
        ws['D14'].alignment = c_c_alignment
        ws['D14'] = '中位数'
        ws['D14'].border = f_border

        ws['E14'].font = title_font
        ws['E14'].alignment = c_c_alignment
        ws['E14'] = '最小值'
        ws['E14'].border = f_border

        ws['F14'].font = title_font
        ws['F14'].alignment = c_c_alignment
        ws['F14'] = '最大值'
        ws['F14'].border = f_border

        ws['G14'].font = title_font
        ws['G14'].alignment = c_c_alignment
        ws['G14'] = '样本标准差'
        ws['G14'].border = f_border

        ws['H14'].font = title_font
        ws['H14'].alignment = c_c_alignment
        ws['H14'] = '样本方差'
        ws['H14'].border = f_border

        # table_data

        for i, value in enumerate(category):
            row = i + 15
            ws['B' + str(row)].font = name_font
            ws['B' + str(row)].alignment = c_c_alignment
            ws['B' + str(row)] = reporting_period_data['names'][i] + " (" + reporting_period_data['units'][
                i] + "/M²)"
            ws['B' + str(row)].border = f_border

            ws['C' + str(row)].font = name_font
            ws['C' + str(row)].alignment = c_c_alignment
            if reporting_period_data['means_per_unit_area'][i] \
                    or reporting_period_data['means_per_unit_area'][i] == 0:
                ws['C' + str(row)] = round(reporting_period_data['means_per_unit_area'][i], 2)
            ws['C' + str(row)].border = f_border

            ws['D' + str(row)].font = name_font
            ws['D' + str(row)].alignment = c_c_alignment
            if reporting_period_data['medians_per_unit_area'][i] \
                    or reporting_period_data['medians_per_unit_area'][i] == 0:
                ws['D' + str(row)] = round(reporting_period_data['medians_per_unit_area'][i], 2)
            ws['D' + str(row)].border = f_border

            ws['E' + str(row)].font = name_font
            ws['E' + str(row)].alignment = c_c_alignment
            if reporting_period_data['minimums_per_unit_area'][i] \
                    or reporting_period_data['minimums_per_unit_area'][i] == 0:
                ws['E' + str(row)] = round(reporting_period_data['minimums_per_unit_area'][i], 2)
            ws['E' + str(row)].border = f_border

            ws['F' + str(row)].font = name_font
            ws['F' + str(row)].alignment = c_c_alignment
            if reporting_period_data['maximums_per_unit_area'][i] \
                    or reporting_period_data['maximums_per_unit_area'][i] == 0:
                ws['F' + str(row)] = round(reporting_period_data['maximums_per_unit_area'][i], 2)
            ws['F' + str(row)].border = f_border

            ws['G' + str(row)].font = name_font
            ws['G' + str(row)].alignment = c_c_alignment
            if (reporting_period_data['stdevs_per_unit_area'][i]) \
                    or reporting_period_data['stdevs_per_unit_area'][i] == 0:
                ws['G' + str(row)] = round(reporting_period_data['stdevs_per_unit_area'][i], 2)
            ws['G' + str(row)].border = f_border

            ws['H' + str(row)].font = name_font
            ws['H' + str(row)].alignment = c_c_alignment
            if reporting_period_data['variances_per_unit_area'][i] \
                    or reporting_period_data['variances_per_unit_area'][i] == 0:
                ws['H' + str(row)] = round(reporting_period_data['variances_per_unit_area'][i], 2)
            ws['H' + str(row)].border = f_border

    ################################################
    # Fourth:
    # 20: title
    # 21~ 26+ca_len*5-1: line
    # 26+ca_len*5: table title
    # 26+ca_len*5~: table_data
    ################################################
    reporting_period_data = report['reporting_period']
    times = reporting_period_data['timestamps']
    has_detail_data_flag = True
    ca_len = len(report['reporting_period']['names'])
    table_row = 27

    if "timestamps" not in reporting_period_data.keys() or \
            reporting_period_data['timestamps'] is None or \
            len(reporting_period_data['timestamps']) == 0:
        has_detail_data_flag = False

    if has_detail_data_flag:
        ws['B20'].font = title_font
        ws['B20'] = name + ' 报告期消耗'
        ws['B26'].font = title_font
        ws['B26'] = name + ' 详细数据'
        # table

        ws['B' + str(table_row)].fill = table_fill
        ws['B' + str(table_row)].border = f_border
        ws['B' + str(table_row)].alignment = c_c_alignment
        ws['B' + str(table_row)] = '日期时间'
        time = times[0]
        has_data = False
        max_row = 0
        if len(time) > 0:
            has_data = True
            max_row = table_row + len(time)
            # print("max_row", max_row)

        if has_data:
            # time
            time_len = len(time)
            for index in range(0, len(time)):
                col = 'B'
                row = str(table_row + 1 + index)
                # col = chr(ord('B') + i)
                ws[col + row].font = title_font
                ws[col + row].alignment = c_c_alignment
                ws[col + row] = time[index]
                ws[col + row].border = f_border
                if index == time_len - 1:
                    row = str(table_row + 2 + index)
                    ws['B' + row].font = title_font
                    ws['B' + row].alignment = c_c_alignment
                    ws['B' + row] = "小计"
                    ws['B' + row].border = f_border
            # data
            for index in range(0, ca_len):
                # table_title
                col = chr(ord('C') + index)

                ws[col + str(table_row)].fill = table_fill
                ws[col + str(table_row)].font = title_font
                ws[col + str(table_row)].alignment = c_c_alignment
                ws[col + str(table_row)] = (reporting_period_data['names'][index] + "(" +
                                            reporting_period_data['units'][index] + ")")
                ws[col + str(table_row)].border = f_border

                # data
                time = times[index]
                time_len = len(time)

                for j in range(0, time_len):
                    row = str(table_row + 1 + j)
                    # col = chr(ord('B') + i)
                    ws[col + row].font = title_font
                    ws[col + row].alignment = c_c_alignment
                    ws[col + row] = round(reporting_period_data['values'][index][j], 0)
                    ws[col + row].border = f_border
                    # subtotals
                    if j == time_len - 1:
                        row = str(table_row + 2 + j)
                        ws[col + row].font = title_font
                        ws[col + row].alignment = c_c_alignment
                        ws[col + row] = round(reporting_period_data['subtotals'][index], 0)
                        ws[col + row].border = f_border

                # bar
                bar = BarChart()
                labels = Reference(ws, min_col=2, min_row=table_row + 1, max_row=max_row)
                bar_data = Reference(ws, min_col=3 + index, min_row=table_row, max_row=max_row)  # openpyxl bug
                bar.add_data(bar_data, titles_from_data=True)
                bar.set_categories(labels)
                bar.height = 5.25  # cm 1.05*5 1.05cm = 30 pt
                bar.width = 18
                # pie.title = "Pies sold by category"
                bar.dLbls = DataLabelList()
                # bar.dLbls.showCatName = True  # label show
                bar.dLbls.showVal = True  # val show
                bar.dLbls.showPercent = True  # percent show
                # s1 = CharacterProperties(sz=1800)     # font size *100
                chart_col = 'B'
                chart_cell = chart_col + str(21 + 5 * index)
                ws.add_chart(bar, chart_cell)

    # ################################################
    # # Fourth: 相关参数
    # # table_row+2: title
    # # 21~ 26+ca_len*5-1: LineChart
    # # 26+ca_len*5: table title
    # # 26+ca_len*5~: table_data
    # ################################################
    #
    # reporting_period_data = report['parameters']
    # times = reporting_period_data['timestamps']
    # has_detail_data_flag = True
    # ca_len = len(reporting_period_data['names'])

    filename = str(uuid.uuid4()) + '.xlsx'
    wb.save(filename)

    return filename
